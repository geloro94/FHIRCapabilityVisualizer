import re
import openpyxl
import requests
import itertools
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from OperationDefinition import get_operation_definition_for_url, initialize_operation_definitions

SERVER_URLS = [
    "https://hapi.fhir.org/baseR4",
    "https://r4.ontoserver.csiro.au/fhir",
    "https://r4.smarthealthit.org",
    #    "https://blaze.imi.uni-luebeck.de/fhir",
    "http://tx.fhir.org/r4",
    "https://tergi.elga.gv.at/fhir-server/api/v4"
]
PATHS_TO_CHECK = [
    # CodeSystem
    "/CodeSystem/$validate-code",
    "/CodeSystem/$lookup",
    "/CodeSystem/$subsumes",
    "/CodeSystem/$find-matches",

    # ConceptMap
    "/ConceptMap/$translate",
    "/ConceptMap/$closure",

    # ValueSet
    "/ValueSet/$expand",
    "/ValueSet/$validate-code",

    # Closure
    "/$closure",

    "/$expand",
    "/$lookup",
    "/$validate-code",
    "/$translate"
]
OPERATION_DEFINITIONS = initialize_operation_definitions()


def get_operation_definitions(url_dict):
    operation_definitions = {}
    for operation_name, operation_definition_url in url_dict.items():
        operation_definitions[operation_name] = get_operation_definition_for_url(operation_definition_url)
    return operation_definitions


def check_parameter_coverage(server_operation_definition, official_operation_definition):

    # Initialize coverage dict for all parameters allowed by official operation definition
    single_operation_parameter_coverage = {}
    for parameter_use, parameters in official_operation_definition.parameters.items():
        single_operation_parameter_coverage[parameter_use] = {parameter.name: False for parameter in parameters}

    # Check server operation definition parameters
    for parameter in server_operation_definition.get('parameter', []):
        parameter_name = parameter.get('name')
        parameter_use = parameter.get('use')
        if parameter_name is not None and parameter_use is not None:
            parameters_for_use = single_operation_parameter_coverage.get(parameter_use, {})
            if parameter_name in parameters_for_use:
                parameters_for_use[parameter_name] = True

    return single_operation_parameter_coverage


def write_parameter_coverage_to_cells(work_sheet, absolute_row_idx, operation_endpoint_name, parameters, parameter_coverage):
    work_sheet.cell(row=absolute_row_idx + 1, column=1, value=f"Parameters for {operation_endpoint_name}")

    for row_idx, server_url in enumerate(parameter_coverage.keys(), start=2):
        work_sheet.cell(row=absolute_row_idx + row_idx, column=1, value=server_url)

    combined_parameters = itertools.chain.from_iterable(parameters.values())

    for column_idx, parameter in enumerate(combined_parameters, start=2):
        # Parameter name
        cell = work_sheet.cell(row=absolute_row_idx + 1, column=column_idx, value=parameter.name)
        if parameter.use == 'in':
            cell.fill = PatternFill(
                start_color="2596be", end_color="2596be", fill_type="solid"
            )
        else:
            cell.fill = PatternFill(
                start_color="e28743", end_color="e28743", fill_type="solid"
            )

        for row_idx, server_tuple in enumerate(parameter_coverage.items(), start=2):
            # Indication of coverage
            server_url, server_coverage = server_tuple
            cell = work_sheet.cell(row=absolute_row_idx + row_idx, column=column_idx)
            if server_coverage is not None:
                covered = server_coverage[parameter.use][parameter.name]
                cell.value = "Yes" if covered else "No"
                if covered:
                    cell.fill = PatternFill(
                        start_color="22bb45", end_color="22bb45", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="D22B2B", end_color="D22B2B", fill_type="solid"
                    )
            else:
                cell.value = "Could not resolve operation definition"
                cell.fill = PatternFill(
                    start_color="a3a3a3", end_color="a3a3a3", fill_type="solid"
                )


def write_missing_official_operation_definition_fault_to_cells(work_sheet, absolute_row_idx, operation_endpoint_name, parameter_coverage):
    work_sheet.cell(row=absolute_row_idx + 1, column=1, value=f"Parameters for {operation_endpoint_name}")
    cell = work_sheet.cell(row=absolute_row_idx + 1, column=2, value=f"Official operation definition could not be resolved")
    cell.fill = PatternFill(start_color="a3a3a3", end_color="a3a3a3", fill_type="solid")

    for row_idx, server_url in enumerate(parameter_coverage, start=2):
        work_sheet.cell(row=absolute_row_idx + row_idx, column=1, value=server_url)
        cell = work_sheet.cell(row=absolute_row_idx + row_idx, column=2)
        cell.fill = PatternFill(start_color="b8b8b8", end_color="b8b8b8", fill_type="solid")


def get_capability_statement(fhir_base_url):
    url = f"{fhir_base_url}/metadata"
    headers = {"Accept": "application/fhir+json"}
    response = requests.get(url, headers=headers)

    if response.status_code != 200:
        raise Exception(f"Error getting capability statement: {response.status_code}")

    return response.json()


def check_endpoint_support(capability_statement, endpoint_paths):
    rest_resources = capability_statement.get("rest", [])
    supported_endpoints = []

    for endpoint_path in endpoint_paths:
        match = re.match(r"/([^/]+)/\$(.+)", endpoint_path)
        if match:
            resource_type, operation_of_interest = match.groups()
        else:
            resource_type = None
            operation_of_interest = endpoint_path[2:]

        supported = False
        definition = None
        for rest_resource in rest_resources:
            if resource_type is None:
                operation_list = rest_resource.get("operation", [])
            else:
                resource_list = rest_resource.get("resource", [])
                operation_list = [
                    operation
                    for resource in resource_list
                    if resource.get("type") == resource_type
                    for operation in resource.get("operation", [])
                ]

            for operation in operation_list:
                if operation.get("name") == operation_of_interest:
                    supported = True
                    definition = operation.get("definition")
                    break

            if supported:
                break

        supported_endpoints.append((supported, definition))

    return supported_endpoints


def auto_size_columns(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length


def auto_size_rows(ws):
    for row_cells in ws.iter_rows():
        sequence = [cell.value.count("\n") for cell in row_cells if cell.value]
        if len(sequence) > 0:
            height = max(sequence) + 1
            ws.row_dimensions[row_cells[0].row].height = 15 * height


def main():
    ####################################################################################################################
    # Endpoint Support
    ####################################################################################################################

    results = [["Endpoint"] + SERVER_URLS]
    supported_endpoints = {endpoint_path: {server_url: (False, None) for server_url in SERVER_URLS} for endpoint_path in
                           PATHS_TO_CHECK}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Endpoint Support"

    for server_index, server_url in enumerate(SERVER_URLS, start=1):
        for index, endpoint_path in enumerate(PATHS_TO_CHECK, start=1):
            row = [endpoint_path]
            ws.cell(row=index + 1, column=1, value=endpoint_path)

            try:
                capability_statement = get_capability_statement(server_url)
                supported, definition = check_endpoint_support(
                    capability_statement, [endpoint_path]
                )[0]
                row.append("Yes" if supported else "No")

                if supported:
                    supported_endpoints[endpoint_path][server_url] = (True, definition)

                cell = ws.cell(row=index + 1, column=server_index + 1)
                cell.value = "Yes" if supported else "No"

                if supported:
                    cell.fill = PatternFill(
                        start_color="22bb45", end_color="22bb45", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="D22B2B", end_color="D22B2B", fill_type="solid"
                    )

            except Exception as e:
                print(f"Error checking {server_url}: {e}")
                row.append("Error")

        results.append(row)

    for index, server_url in enumerate(SERVER_URLS, start=1):
        ws.cell(row=1, column=index + 1, value=server_url)

    auto_size_columns(ws)
    auto_size_rows(ws)

    ####################################################################################################################
    # Parameter Coverage
    ####################################################################################################################

    ws = wb.create_sheet("Parameter Coverage")

    curr_row_idx = 1
    for endpoint, supported_on_server in supported_endpoints.items():

        if endpoint in OPERATION_DEFINITIONS:
            parameter_coverage_by_server = {server_url: None for server_url in supported_on_server.keys()}

            fhir_operation_definition = OPERATION_DEFINITIONS.get(endpoint)
            if fhir_operation_definition is None:
                # Skip if there is no source provided for the official definition
                print(f"Could not retrieve official operation definition for {endpoint}")
                parameter_coverage_by_server = [server_url for server_url in supported_on_server.keys]
                write_missing_official_operation_definition_fault_to_cells(ws, curr_row_idx, endpoint, parameter_coverage_by_server)
                continue

            for server_url, value in supported_on_server.items():
                supported, definition = value
                if supported:
                    try:
                        operation_definition_json = requests.get(definition, headers={'Accept': 'application/fhir+json'}).json()
                        parameter_coverage = check_parameter_coverage(operation_definition_json, fhir_operation_definition)
                        parameter_coverage_by_server[server_url] = parameter_coverage
                    except Exception as exc:
                        print(f"Could not process operation {endpoint} for server {server_url} ({definition}): \n {exc}")

            parameters = OPERATION_DEFINITIONS[endpoint].parameters
            write_parameter_coverage_to_cells(ws, curr_row_idx, endpoint, parameters, parameter_coverage_by_server)
            curr_row_idx += len(parameter_coverage_by_server.keys()) + 2

    auto_size_columns(ws)
    auto_size_rows(ws)

    ####################################################################################################################

    wb.save("capabilities_visualized.xlsx")

    print("Results saved to capabilities_visualized.xlsx")


if __name__ == "__main__":
    main()
